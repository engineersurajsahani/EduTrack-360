import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from django.utils import timezone

def export_students_to_excel(students):
    """Exports student progress records to a styled Excel worksheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Progress Report"

    # Header Title
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = "EDUTRACK 360 - SMART STUDENT PROGRESS & ACHIEVEMENT REPORT"
    title_cell.font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtitle
    ws.merge_cells('A2:L2')
    sub_cell = ws['A2']
    sub_cell.value = f"Generated on {timezone.now().strftime('%d %B %Y, %I:%M %p')} | Confidential College Record"
    sub_cell.font = Font(name='Calibri', size=10, italic=True, color="FFFFFF")
    sub_cell.fill = PatternFill(start_color="283593", end_color="283593", fill_type="solid")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Table Column Headers
    headers = [
        "PRN", "Roll No", "Student Name", "Branch", "Semester", 
        "Attendance %", "Academic %", "NSS %", "Cultural %", "Sports %", 
        "Overall Score %", "Progress Status"
    ]
    
    ws.append([]) # Row 3 blank
    ws.append(headers) # Row 4 headers
    ws.row_dimensions[4].height = 25

    header_fill = PatternFill(start_color="3949AB", end_color="3949AB", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Data Rows
    row_idx = 5
    for student in students:
        overall = student.get_overall_progress()
        acad = student.get_academic_percentage()
        nss = student.get_specific_category_pct('NSS')
        cult = student.get_specific_category_pct('CULTURAL')
        sports = student.get_specific_category_pct('SPORTS')
        status = "EXCELLENT (Green)" if overall >= 75 else ("AVERAGE (Yellow)" if overall >= 50 else "CRITICAL (Red)")

        row_data = [
            student.prn,
            student.roll_no,
            student.user.get_full_name() or student.user.username,
            student.branch.code,
            student.semester.roman_name,
            f"{student.attendance_percentage}%",
            f"{acad}%",
            f"{nss}%",
            f"{cult}%",
            f"{sports}%",
            f"{overall}%",
            status
        ]
        ws.append(row_data)
        ws.row_dimensions[row_idx].height = 20

        # Row styling
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_num not in [3] else "left", vertical="center")
            
            # Status highlight
            if col_num == 12:
                if overall >= 75:
                    cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                    cell.font = Font(color="2E7D32", bold=True)
                elif overall >= 50:
                    cell.fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
                    cell.font = Font(color="F57F17", bold=True)
                else:
                    cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                    cell.font = Font(color="C62828", bold=True)

        row_idx += 1

    # Auto adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=EduTrack360_Progress_Report_{timezone.now().strftime("%Y%m%d")}.xlsx'
    return response
